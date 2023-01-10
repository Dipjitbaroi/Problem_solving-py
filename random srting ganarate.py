import openpyxl as op
import strgen as sg
# strgen_res = sg.StringGenerator("[\w]{3}").render()
# strgen_res1 = sg.StringGenerator("[\d]{2}").render()
# print(strgen_res)
wb = op.load_workbook('Books.xlsx')
ws = wb.active
for i in range(2,1000001):
    ws.cell(row=i,column=1).value = sg.StringGenerator("[\w]{3}").render()
    ws.cell(row=i,column=2).value = sg.StringGenerator("[\d]{2}").render()
wb.save('Books1.xlsx')