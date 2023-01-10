import openpyxl
wb = openpyxl.load_workbook('Book1.xlsx')
names = []
marks = []
ws = wb.active
for columns in ws.iter_cols(1, 1):
   for cell in columns:
       if cell.value != "Names":
           names.append(cell.value)
     # print (cell.value)
for columns in ws.iter_cols(2):
    for cell in columns:
        if cell.value != "Marks":
            marks.append(cell.value)
for i in range(41,80):
    for name in names:
        ws.cell(row=i,column=1).value = name
    for mark in marks:
        ws.cell(row=i,column=2).value = mark

# print(ws.cell(row=41,column=1).value)
# ws['A40'] = 'abal'
# print(ws["A41"])
wb.save('Book2.xlsx')

# for rows in range (41,80):
#     ws.cell(row=rows,column=1).value = names
print(names)
# print(marks)
# sheet1 = xl("Marks")
# print(sheet1)
# rangeselected = []
# for i in range(1,40,1):
#     rangeselected.append(sheet1.cell(row=i,column=2).value)
#     # rangeselected.append(sheet1.cell(row=i,column=2).value)
# print(rangeselected)